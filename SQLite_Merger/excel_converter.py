import csv
import os
from pathlib import Path

from openpyxl import load_workbook

from constants import DEFAULT_CODEC
from utils import get_valid_encoding


class ExcelConverter:
    """Conversion tables Excel en fichiers CSV"""

    def tables_to_csv(self, xl_file: Path, tables_mapping: dict[str, tuple[Path, str]]):
        """Crée des CSV à partir du fichier Excel.
        La clé du dictionnaire doit être le nom de la table excel
        et la valeur un tuple indiquant le fichier à créer et l'encodage à utiliser."""
        for table_name, (csv_file, encoding) in tables_mapping.items():
            self.table_to_csv(xl_file, table_name, str(csv_file), encoding)

    def table_to_csv(self, xl_file: Path, table_name: str, csv_filename: str | Path, encoding: str):
        """Exporte une table Excel vers un fichier CSV"""
        data = self._table_to_list(xl_file, table_name)
        valid_encoding = encoding if get_valid_encoding(encoding) else DEFAULT_CODEC

        if not data:
            # Créer un fichier CSV vide avec juste l'en-tête si pas de données
            with open(csv_filename, "w", encoding=valid_encoding, newline="") as f:
                f.flush()  # vide le buffer Python
                os.fsync(f.fileno())  # force l'écriture sur disque

            return

        with open(csv_filename, "w", encoding=valid_encoding, newline="") as f:
            writer = csv.writer(f, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerows(data)

    def _table_to_list(self, xl_file: Path, table_name: str) -> list[list]:
        """Lit une table Excel et retourne un tableau 2D"""
        try:
            wb = None  # pour éviter une erreur dans le finally
            wb = load_workbook(xl_file, data_only=True)
            for ws in wb.worksheets:
                if table_name in ws.tables:
                    table = ws.tables[table_name]
                    data = [[cell.value for cell in row] for row in ws[table.ref]]
                    return data
            raise KeyError(f"Table '{table_name}' non trouvée")
        except Exception as e:
            raise Exception(f"Erreur inattendue table Excel '{table_name}'") from e
        finally:
            if wb:
                wb.close()
